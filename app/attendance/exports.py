from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.jalali import jalali_month_range, jalali_day_to_gregorian, get_weekday_names_from_jalali
from attendance.models import AttendanceDay

STATUS_CODE = {
    "ABSENT": "غیر حاضر",
    "SHIFT_OFF": "نوبتی",
    "HOLIDAY": "رخصتی",
    "LEAVE": "رخصت",
}


def build_attendance_xlsx(jy: int, jm: int, employees):
    rng = jalali_month_range(jy, jm)
    days = list(range(1, rng.days + 1))

    friday_days = set()
    for d in days:
        g = jalali_day_to_gregorian(jy, jm, d)
        if g.weekday() == 4:  # Friday
            friday_days.add(d)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{jy}-{jm:02d}"

    headers = [
        "Employee ID",
        "First Name",
        "Father Name",
    ] + [f"{jy}-{jm}-{d} {get_weekday_names_from_jalali(jy, jm, d)[1]}" for d in days]
    ws.append(headers)

    exceptions = AttendanceDay.objects.filter(
        employee__in=employees,
        date__range=(rng.g_start, rng.g_end),
    ).values("employee_id", "date", "status")

    exc_map = {(e["employee_id"], e["date"]): e["status"] for e in exceptions}

    for emp in employees:
        row = [emp.id, emp.first_name, emp.father_name]
        for d in days:
            g_date = jalali_day_to_gregorian(jy, jm, d)

            cell = "جمعه" if d in friday_days else "حاضر"
            st = exc_map.get((emp.id, g_date))
            if st:
                cell = STATUS_CODE.get(st, st)

            row.append(cell)

        ws.append(row)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    for col in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    return wb
