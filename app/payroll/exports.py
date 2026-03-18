from decimal import Decimal, ROUND_CEILING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from attendance.models import AttendanceDay
from core.jalali import (
    get_weekday_names_from_jalali,
    jalali_day_to_gregorian,
    jalali_month_range,
)
from overtime.models import OvertimeEntry


STATUS_CODE = {
    "ABSENT": "غیر حاضر",
    "SHIFT_OFF": "نوبتی",
    "HOLIDAY": "رخصتی",
    "LEAVE": "رخصت",
}

THIN = Side(style="thin", color="000000")
HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
SUBHEADER_FILL = PatternFill("solid", fgColor="F3F3F3")
CENTER = Alignment(horizontal="center", vertical="center")
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def dceil2(x):
    if x is None:
        return Decimal("0.00")
    return Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_CEILING)


def _ordered_lines(run, order_by: str = "name"):
    lines = run.lines.select_related("employee", "employee__department", "employee__position")
    if order_by == "id":
        return list(lines.order_by("employee__id", "employee__first_name", "employee__father_name"))
    return list(lines.order_by("employee__first_name", "employee__father_name", "employee__id"))


def _set_base_employee_column_widths(ws):
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22


def _build_attendance_sheet(wb, run, lines):
    jy, jm = run.year, run.month
    rng = jalali_month_range(jy, jm)
    days = list(range(1, rng.days + 1))

    friday_days = set()
    for d in days:
        g = jalali_day_to_gregorian(jy, jm, d)
        if g.weekday() == 4:  # Friday
            friday_days.add(d)

    ws = wb.active
    ws.title = "Attendance"

    headers = [
        "Employee ID",
        "First Name",
        "Father Name",
    ] + [
        f"{jy}-{jm}-{d} {get_weekday_names_from_jalali(jy, jm, d)[1]}" for d in days
    ] + [
        "Leave Count",
        "Absent Count",
        "Present Count",
    ]
    ws.append(headers)

    employee_ids = [line.employee_id for line in lines]
    exceptions = AttendanceDay.objects.filter(
        employee_id__in=employee_ids,
        date__range=(rng.g_start, rng.g_end),
    ).values("employee_id", "date", "status")
    exc_map = {(e["employee_id"], e["date"]): e["status"] for e in exceptions}

    for line in lines:
        emp = line.employee
        row = [emp.id, emp.first_name, emp.father_name]
        leave_count = 0
        absent_count = 0
        present_count = 0

        for d in days:
            g_date = jalali_day_to_gregorian(jy, jm, d)
            status = exc_map.get((emp.id, g_date))

            cell = "جمعه" if d in friday_days else "حاضر"
            if status:
                cell = STATUS_CODE.get(status, status)

            if status == AttendanceDay.Status.LEAVE:
                leave_count += 1
            elif status == AttendanceDay.Status.ABSENT:
                absent_count += 1
            elif not status and d not in friday_days:
                present_count += 1
            elif status == AttendanceDay.Status.PRESENT:
                present_count += 1

            row.append(cell)

        row.extend([leave_count, absent_count, present_count])
        ws.append(row)

    _set_base_employee_column_widths(ws)
    for col in range(4, 4 + len(days)):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions[get_column_letter(4 + len(days))].width = 14
    ws.column_dimensions[get_column_letter(5 + len(days))].width = 14
    ws.column_dimensions[get_column_letter(6 + len(days))].width = 14



def _build_overtime_sheet(wb, run, lines):
    jy, jm = run.year, run.month
    rng = jalali_month_range(jy, jm)
    days = list(range(1, rng.days + 1))

    ws = wb.create_sheet("Overtime")
    headers = [
        "Employee ID",
        "First Name",
        "Father Name",
    ] + [str(d) for d in days] + [
        "Total Hours",
        "Total Overtime Payment",
    ]
    ws.append(headers)

    employee_ids = [line.employee_id for line in lines]
    entries = OvertimeEntry.objects.filter(
        employee_id__in=employee_ids,
        date__range=(rng.g_start, rng.g_end),
    ).values("employee_id", "date", "hours")
    hours_map = {(e["employee_id"], e["date"]): e["hours"] for e in entries}
    overtime_amount_map = {line.employee_id: dceil2(line.overtime) for line in lines}

    for line in lines:
        emp = line.employee
        row = [emp.id, emp.first_name, emp.father_name]
        total_hours = Decimal("0.00")

        for d in days:
            g_date = jalali_day_to_gregorian(jy, jm, d)
            hours = dceil2(hours_map.get((emp.id, g_date), 0))
            total_hours += hours
            row.append(float(hours))

        row.append(float(dceil2(total_hours)))
        row.append(float(overtime_amount_map.get(emp.id, Decimal("0.00"))))
        ws.append(row)

    _set_base_employee_column_widths(ws)
    for col in range(4, 4 + len(days)):
        ws.column_dimensions[get_column_letter(col)].width = 9
    ws.column_dimensions[get_column_letter(4 + len(days))].width = 14
    ws.column_dimensions[get_column_letter(5 + len(days))].width = 22



def _build_payroll_sheet(wb, lines):
    ws = wb.create_sheet("Payroll")

    headers = [
        "Employee ID",
        "First Name",
        "Father Name",
        "Base Salary",
        "Attendance Deduction",
        "Salary",
        "Bonus",
        "Overtime",
        "Total",
        "Tax",
        "Prepaid",
        "Amount To Pay",
    ]
    ws.append(headers)

    for line in lines:
        ws.append([
            line.employee.id,
            line.employee.first_name,
            line.employee.father_name,
            float(dceil2(line.base_salary)),
            float(dceil2(line.attendance_deduction)),
            float(dceil2(line.salary)),
            float(dceil2(line.bonus)),
            float(dceil2(line.overtime)),
            float(dceil2(line.total)),
            float(dceil2(line.tax)),
            float(dceil2(line.prepaid)),
            float(dceil2(line.amount_to_pay)),
        ])

    def sum_field(name):
        return sum((dceil2(getattr(line, name)) for line in lines), Decimal("0.00"))

    ws.append([
        "TOTALS",
        "",
        "",
        float(sum_field("base_salary")),
        float(sum_field("attendance_deduction")),
        float(sum_field("salary")),
        float(sum_field("bonus")),
        float(sum_field("overtime")),
        float(sum_field("total")),
        float(sum_field("tax")),
        float(sum_field("prepaid")),
        float(sum_field("amount_to_pay")),
    ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    _set_base_employee_column_widths(ws)



def _style_summary_sheet(ws, last_row: int):
    for row in range(1, last_row + 1):
        for col in range(1, 17):
            cell = ws.cell(row, col)
            cell.alignment = WRAP_CENTER
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if row in (1, 2):
                cell.font = Font(bold=True)

    for merged in ["B1:D1", "F1:H1", "I1:K1", "E1:E2", "L1:L2", "M1:M2", "N1:O1", "P1:P2"]:
        ws.merge_cells(merged)

    for coord in ["A1", "B1", "F1", "I1", "E1", "L1", "M1", "N1", "P1"]:
        ws[coord].fill = HEADER_FILL
    for coord in ["B2", "C2", "D2", "G2", "H2", "I2", "J2", "K2", "N2", "O2"]:
        ws[coord].fill = SUBHEADER_FILL

    widths = {
        "A": 12, "B": 12, "C": 12, "D": 12, "E": 14, "F": 14, "G": 12, "H": 12,
        "I": 14, "J": 14, "K": 14, "L": 14, "M": 32, "N": 22, "O": 22, "P": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 24



def _build_format_1_sheet(wb, lines):
    ws = wb.create_sheet("1")

    ws["A1"] = "رسیدات"
    ws["B1"] = "ملاحظات"
    ws["E1"] = "قابل تادیه"
    ws["F1"] = "وضع شده"
    ws["I1"] = "استحقاق تنخوا و اضافه كاري"
    ws["L1"] = "اصل معاش"
    ws["M1"] = "Department - Position"
    ws["N1"] = "شـــــهــرت"
    ws["P1"] = "شماره"

    ws["B2"] = "رخصت"
    ws["C2"] = "غیرحاضر"
    ws["D2"] = "حاضر"
    ws["G2"] = "پيشکى"
    ws["H2"] = "ماليه"
    ws["I2"] = "مجموع"
    ws["J2"] = "اضافه کارى"
    ws["K2"] = "معاش"
    ws["N2"] = "ولد"
    ws["O2"] = "اسم"

    attendance_ws = "Attendance"
    payroll_ws = "Payroll"

    for idx, line in enumerate(lines, start=3):
        payroll_row = idx - 1
        attendance_row = idx - 1
        emp = line.employee
        dept_position = f"{emp.department.name} - {emp.position.name}" if emp.department_id and emp.position_id else ""

        ws[f"B{idx}"] = f"={attendance_ws}!AG{attendance_row}"
        ws[f"C{idx}"] = f"={attendance_ws}!AH{attendance_row}"
        ws[f"D{idx}"] = f"={attendance_ws}!AI{attendance_row}"
        ws[f"E{idx}"] = f"=K{idx}-F{idx}"
        ws[f"F{idx}"] = f"=H{idx}+G{idx}"
        ws[f"G{idx}"] = f"={payroll_ws}!K{payroll_row}"
        ws[f"H{idx}"] = f"={payroll_ws}!J{payroll_row}"
        ws[f"I{idx}"] = f"=K{idx}+J{idx}"
        ws[f"J{idx}"] = f"={payroll_ws}!H{payroll_row}"
        ws[f"K{idx}"] = f"={payroll_ws}!F{payroll_row}"
        ws[f"L{idx}"] = f"={payroll_ws}!D{payroll_row}"
        ws[f"M{idx}"] = dept_position
        ws[f"N{idx}"] = f"={attendance_ws}!C{attendance_row}"
        ws[f"O{idx}"] = f"={attendance_ws}!B{attendance_row}"
        ws[f"P{idx}"] = emp.id

    total_row = len(lines) + 3
    if lines:
        for col in "BCDEFGHIJKL":
            ws[f"{col}{total_row}"] = f"=SUM({col}3:{col}{total_row - 1})"
        ws[f"M{total_row}"] = "TOTALS"
        ws[f"P{total_row}"] = len(lines)

    _style_summary_sheet(ws, total_row)



def build_payroll_xlsx(run, order_by: str = "name"):
    wb = Workbook()
    lines = _ordered_lines(run, order_by=order_by)

    _build_attendance_sheet(wb, run, lines)
    _build_overtime_sheet(wb, run, lines)
    _build_payroll_sheet(wb, lines)
    _build_format_1_sheet(wb, lines)

    return wb
