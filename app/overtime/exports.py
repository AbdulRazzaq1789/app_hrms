from decimal import Decimal, ROUND_CEILING
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.jalali import jalali_month_range, jalali_day_to_gregorian
from overtime.models import OvertimeEntry


def ceil2(x):
    if x is None:
        return Decimal("0.00")
    return Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_CEILING)


def build_overtime_xlsx(jy: int, jm: int, employees):
    rng = jalali_month_range(jy, jm)
    days = list(range(1, rng.days + 1))

    wb = Workbook()
    ws = wb.active
    ws.title = f"{jy}-{jm:02d}"

    headers = ["Employee ID", "First Name", "Father Name"] + [str(d) for d in days] + ["Total Hours"]
    ws.append(headers)

    entries = OvertimeEntry.objects.filter(
        employee__in=employees,
        date__range=(rng.g_start, rng.g_end),
    ).values("employee_id", "date", "hours")

    hours_map = {(e["employee_id"], e["date"]): e["hours"] for e in entries}

    for emp in employees:
        row = [emp.id, emp.first_name, emp.father_name]
        total = Decimal("0.00")

        for d in days:
            g_date = jalali_day_to_gregorian(jy, jm, d)
            h = ceil2(hours_map.get((emp.id, g_date), 0))
            total += h
            row.append(float(h))

        row.append(float(ceil2(total)))
        ws.append(row)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    for col in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 9

    return wb
